# data/past-lives.json -> data/past-lives.xlsx 변환 스크립트
# 실행: python scripts/make_xlsx.py
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
records = json.loads((ROOT / "data" / "past-lives.json").read_text(encoding="utf-8"))

wb = Workbook()
ws = wb.active
ws.title = "전생 기록 150건"

HEADERS = [
    ("번호", 7),
    ("전생의 직업 · 존재", 34),
    ("시대", 34),
    ("사인 (죽은 이유)", 56),
    ("전생의 업적", 56),
    ("사람들의 기억", 56),
]

header_font = Font(name="Malgun Gothic", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4A3B8C")
body_font = Font(name="Malgun Gothic", size=10)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top")

for col, (title, width) in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width

for i, r in enumerate(records, start=2):
    values = [r["id"], r["being"], r["era"], r["death"], r["achievement"], r["memory"]]
    for col, v in enumerate(values, start=1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.font = body_font
        cell.border = border
        cell.alignment = center if col == 1 else wrap

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(records) + 1}"

out = ROOT / "data" / "past-lives.xlsx"
wb.save(out)
print(f"저장 완료: {out} ({len(records)}건)")
